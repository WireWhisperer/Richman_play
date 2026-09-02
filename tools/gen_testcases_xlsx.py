# -*- coding: utf-8 -*-
"""将 Group3_Testcases.json（v2.0）的全部用例导出为一张中文 xlsx 总览表（每行一个用例，按 id 排序）。"""
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

def find_repo_root():
    """定位仓库根目录：脚本位于 tools/ 内时取上一级；否则取同级 Richman_play 目录。"""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    parent = os.path.dirname(script_dir)
    if os.path.isdir(os.path.join(parent, "testcases")) and \
       os.path.isdir(os.path.join(parent, "src")):
        return parent
    return os.path.join(script_dir, "Richman_play")

SRC = os.path.join(find_repo_root(), "testcases", "Group3_Testcases.json")
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "测试用例汇总表.xlsx")

ID_NAME = {"Q": "钱夫人(Q)", "A": "阿土伯(A)", "S": "孙小美(S)", "J": "金贝贝(J)"}
STATUS_NAME = {"NORMAL": "正常", "BANKRUPT": "破产"}
PHASE_NAME = {"COMMAND": "命令阶段", "PROMPT": "提示回答阶段", "ENDED": "已结束"}
ERR_NAME = {
    "INVALID_JSON": "JSON 非法",
    "UNSUPPORTED_VERSION": "不支持的 schema_version",
    "UNSUPPORTED_MODE": "不支持的 mode",
    "INVALID_PRESET": "前置状态非法",
    "INVALID_MAP": "地图错误",
    "INVALID_COMMAND": "不支持的命令",
    "INVALID_PARAMS": "参数非法",
    "INVALID_PHASE": "当前阶段不允许该命令",
    "DICE_SEQUENCE_EMPTY": "骰子序列耗尽",
    "RANDOM_SEQUENCE_EMPTY": "随机流耗尽",
    "RANDOM_VALUE_OUT_OF_RANGE": "随机流值越界",
    "ACTION_AFTER_END": "游戏结束后仍有操作",
}
FIELD_CN = {"fund": "资金", "credit": "点数", "position": "位置",
            "status": "状态", "god_of_wealth_rounds": "财神回合",
            "turn_number": "回合号"}


def item_cn(kind):
    return {"BLOCK": "路障", "ROBOT": "机器娃娃"}.get(kind, kind)


def describe_player(p):
    parts = [f"{ID_NAME.get(p['id'], p['id'])}：资金{p['fund']}，点数{p['credit']}，位置{p['position']}"]
    st = p.get("status", "NORMAL")
    if st != "NORMAL":
        parts.append(STATUS_NAME.get(st, st))
    it = p.get("items", {})
    has = [f"{item_cn(k)}×{v}" for k, v in it.items() if v]
    if has:
        parts.append("道具：" + "、".join(has))
    god = p.get("god_of_wealth_rounds", 0)
    if god:
        parts.append(f"财神{god}回合")
    return "，".join(parts)


def describe_fortune(f):
    if not f:
        return None
    pos = f.get("position")
    if pos is None:
        s = "地图财神：无"
    else:
        s = f"地图财神：位置{pos}，保留{f.get('remaining_map_turns')}回合"
    if f.get("next_spawn_after_turn") is not None:
        s += f"，第{f['next_spawn_after_turn']}回合完成时生成"
    return s


def describe_preset(preset):
    lines = []
    lines.append(f"回合号：{preset.get('turn_number', 1)}")
    cur = preset.get("current_user")
    if cur:
        lines.append(f"当前玩家：{ID_NAME.get(cur, cur)}")
    for p in preset.get("players", []):
        lines.append(describe_player(p))
    props = preset.get("properties") or []
    if props:
        lines.append("地产：" + "；".join(
            f"位置{p['position']}属{ID_NAME.get(p['owner'], p['owner'])}等级{p['level']}"
            for p in props))
    mis = preset.get("map_items") or []
    if mis:
        lines.append("地图道具：" + "；".join(
            f"位置{m['position']}有{item_cn(m['type'])}" for m in mis))
    f = describe_fortune(preset.get("fortune"))
    if f:
        lines.append(f)
    rng = preset.get("random_control")
    if rng:
        mode = rng.get("mode")
        if mode == "SEQUENCE":
            streams = {k: list(v) for k, v in (rng.get("streams") or {}).items()}
            if streams:
                lines.append("随机流：" + "；".join(f"{k}={v}" for k, v in streams.items()))
        else:
            lines.append(f"随机流：{mode} seeds={rng.get('stream_seeds')}")
    if preset.get("dice_sequence") is not None:
        lines.append(f"骰子序列：{preset['dice_sequence']}")
    return "\n".join(lines)


def describe_action(a):
    cmd = (a.get("command") or "").upper()
    params = a.get("params") or {}
    if cmd == "ROLL":
        return "掷骰子（ROLL）"
    if cmd == "STEP":
        return f"移动 {params.get('steps')} 步（STEP）"
    if cmd == "SELL":
        return f"出售位置 {params.get('position')} 的地产（SELL）"
    if cmd == "BLOCK":
        return f"在偏移 {params.get('offset')} 处放置路障（BLOCK）"
    if cmd == "ROBOT":
        return "使用机器娃娃清除前方10格路障（ROBOT）"
    if cmd == "QUERY":
        return "查询资产（QUERY）"
    if cmd == "HELP":
        return "查看帮助（HELP）"
    if cmd == "QUIT":
        return "退出游戏（QUIT）"
    if cmd == "ADVANCE_TURN":
        return "原地推进一回合（ADVANCE_TURN）"
    if cmd == "ANSWER":
        v = str(params.get("value", "")).upper()
        if v in ("Y", "N"):
            return f"对提示回答 {v}（ANSWER）"
        return f"选择 {v}（ANSWER）"
    return a.get("command", "")


def describe_expected(exp, eo, ee):
    if eo == "ERROR":
        code = (ee or {}).get("code", "")
        ai = (ee or {}).get("action_index")
        path = (ee or {}).get("path", "")
        s = f"预期产生错误：{code}（{ERR_NAME.get(code, '')}）"
        if ai is not None:
            s += f"，第{ai}个Action"
        if path:
            s += f"，路径{path}"
        return s
    lines = []
    if "turn_number" in exp:
        lines.append(f"回合号：{exp['turn_number']}")
    if "current_user" in exp:
        lines.append(f"当前玩家：{ID_NAME.get(exp['current_user'], exp['current_user'])}")
    if "phase" in exp:
        lines.append(f"阶段：{PHASE_NAME.get(exp['phase'], exp['phase'])}")
    if "game_status" in exp:
        lines.append("游戏状态：已结束" if exp["game_status"] == "FINISHED" else "游戏状态：进行中")
    if "winner" in exp:
        w = exp["winner"]
        lines.append("胜者：" + (ID_NAME.get(w, str(w)) if w else "无"))
    if "pending_prompt" in exp:
        lines.append("待处理提示：无" if exp["pending_prompt"] is None
                     else f"待处理提示：{exp['pending_prompt']}")
    for p in exp.get("players") or []:
        prefix = ID_NAME.get(p.get("id"), p.get("id"))
        fields = []
        for k, v in p.items():
            if k == "id":
                continue
            if k == "status":
                fields.append(f"状态={STATUS_NAME.get(v, v)}")
            elif k == "items":
                its = [f"{item_cn(kk)}×{vv}" for kk, vv in v.items() if vv]
                if its:
                    fields.append("道具=" + "、".join(its))
            elif k == "fields_absent":
                if v:
                    fields.append("不存在字段=" + "、".join(v))
            else:
                fields.append(f"{FIELD_CN.get(k, k)}={v}")
        lines.append(prefix + ("：" + "，".join(fields) if fields else ""))
    for prop in exp.get("properties") or []:
        lines.append(f"地产：位置{prop['position']}属{ID_NAME.get(prop['owner'], prop['owner'])}"
                     f"等级{prop['level']}")
    if "properties_absent" in exp:
        lines.append("无地产位置：" + "、".join(str(x) for x in exp["properties_absent"]))
    for m in exp.get("map_items") or []:
        lines.append(f"地图道具：位置{m['position']}有{item_cn(m['type'])}")
    if "map_items_absent" in exp:
        lines.append("无道具位置：" + "、".join(str(x) for x in exp["map_items_absent"]))
    for d in exp.get("display_players") or []:
        lines.append(f"地图显示：位置{d['position']}显示{d['visible_user']}")
    for d in exp.get("display_cells") or []:
        lines.append(f"格子显示：位置{d['position']}（{d.get('base_type')}）显示"
                     f"{d.get('visible_symbol')}（{d.get('visible_entity')}）")
    if "fortune" in exp:
        f = exp["fortune"]
        if f.get("position") is not None:
            lines.append(f"地图财神：位置{f['position']}，保留{f.get('remaining_map_turns')}回合")
        else:
            lines.append("地图财神：无"
                         + (f"，第{f['next_spawn_after_turn']}回合完成时生成"
                            if f.get("next_spawn_after_turn") is not None else ""))
    if "fortune_assert" in exp:
        lines.append("财神断言：" + json.dumps(exp["fortune_assert"], ensure_ascii=False))
    return "\n".join(lines)


def main():
    with open(SRC, encoding="utf-8") as fp:
        data = json.load(fp)

    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"
    headers = ["用例ID", "用例名称", "前置条件", "操作步骤", "预期结果"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for t in sorted(data["tests"], key=lambda t: t["case_id"]):
        pre = describe_preset(t["preset"])
        acts = t.get("actions") or []
        ops = "\n".join(f"{i + 1}. {describe_action(a)}" for i, a in enumerate(acts)) \
            if acts else "（无操作）"
        exp = describe_expected(t.get("expected", {}),
                                t.get("expected_outcome", "SUCCESS"),
                                t.get("expected_error"))
        ws.append([t["case_id"], t["case_name"], pre, ops, exp])

    widths = [20, 34, 66, 40, 66]
    for col, w in zip("ABCDE", widths):
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    wb.save(OUT)
    print(f"OK: {len(data['tests'])} 个用例 -> {OUT}")


if __name__ == "__main__":
    main()
